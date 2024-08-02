from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
modra_barva = PatternFill("solid", fgColor="0000FFFF")
zluta_barva = PatternFill("solid", fgColor="00FFFF99")

dny = ['PO','UT','ST','CT','PA']
radek = 1
sloupec = 1
for den in dny:
  bunka = ws1.cell(radek, sloupec)
  bunka.value = den
  bunka.fill = zluta_barva
  radek += 1

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
radek = 1
for den in rozvrh_hodin:
  sloupec = 2
  for predmet in den:
    bunka = ws1.cell(radek, sloupec)
    bunka.value = predmet
    if predmet == "Oběd":
      bunka.fill = sediva_barva
    else:
      bunka.fill = modra_barva
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")
