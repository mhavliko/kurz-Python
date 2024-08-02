from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"
cislo_radku = 1
cislo_sloupce = 2

bunka = ws1.cell(cislo_radku, cislo_sloupce)
bunka.value = "Test"

# Určím si barvu
sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
# Přiřadím barvu buňce
bunka.fill = sediva_barva

wb.save(filename="rozvrh.xlsx")